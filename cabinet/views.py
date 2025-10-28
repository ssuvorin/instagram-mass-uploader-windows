from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
import json
from django.utils import timezone
from django.db.models.functions import TruncDate
from django.db.models import Sum
from openpyxl import Workbook
from django.http import HttpResponse
import csv
from django.views.decorators.csrf import csrf_exempt
from .models import Agency, Client, ClientHashtag, CalculationHistory
from django.utils.text import slugify
from uploader.models import HashtagAnalytics
from uploader.models import InstagramAccount, AccountAnalytics
from .forms import AgencyForm, ClientForm, ClientHashtagForm
from django.contrib.auth import get_user_model
from .services import AnalyticsService
from .services import AgencyCalculatorService
from .currency_service import currency_service


@login_required
def dashboard(request):
    if request.user.is_superuser:
        agencies = Agency.objects.select_related("owner").all()
        clients = Client.objects.select_related("agency", "user").all()
        return render(request, "cabinet/admin_dashboard.html", {"agencies": agencies, "clients": clients})

    # Non-superusers must be redirected to their specific cabinet only
    client = Client.objects.filter(user=request.user).select_related("agency").first()
    if client:
        # Redirect client to their personal dashboard
        return redirect("cabinet_client_dashboard")

    agency = Agency.objects.filter(owner=request.user).first()
    if agency:
        return redirect("cabinet_agency_dashboard")

    messages.info(request, "No cabinet role assigned.")
    return redirect("/")


@login_required
def manage_agencies(request):
    if not request.user.is_superuser:
        return redirect("/")
    if request.method == "POST":
        form = AgencyForm(request.POST)
        if form.is_valid():
            # Create dedicated owner user for the agency (never use current admin)
            from django.contrib.auth import get_user_model
            import secrets
            User = get_user_model()
            agency_name = form.cleaned_data.get("name") or "agency"
            base_username = slugify(agency_name) or "agency"
            username = base_username
            suffix = 1
            while User.objects.filter(username=username).exists():
                suffix += 1
                username = f"{base_username}{suffix}"
            password = secrets.token_urlsafe(9)

            owner_user = User.objects.create_user(username=username, password=password)

            agency = form.save(commit=False)
            agency.owner = owner_user
            agency.save()

            messages.success(
                request,
                f"Agency created. Credentials — Login: {username} | Password: {password}"
            )
            return redirect("cabinet_manage_agencies")
    else:
        form = AgencyForm()
    agencies = Agency.objects.select_related("owner").all()
    return render(request, "cabinet/manage_agencies.html", {"form": form, "agencies": agencies})


@login_required
def manage_clients(request):
    if not request.user.is_superuser:
        return redirect("/")
    if request.method == "POST":
        form = ClientForm(request.POST)
        if form.is_valid():
            # Create client and dedicated user for login
            client: Client = form.save(commit=False)
            client.save()

            from django.contrib.auth import get_user_model
            import secrets
            User = get_user_model()
            client_name = client.name or "client"
            base_username = slugify(client_name) or "client"
            username = base_username
            suffix = 1
            while User.objects.filter(username=username).exists():
                suffix += 1
                username = f"{base_username}{suffix}"
            password = secrets.token_urlsafe(9)

            login_user = User.objects.create_user(username=username, password=password)
            client.user = login_user
            client.save(update_fields=["user"])

            messages.success(
                request,
                f"Client created. Credentials — Login: {username} | Password: {password}"
            )
            return redirect("cabinet_manage_clients")
    else:
        form = ClientForm()
    clients = Client.objects.select_related("agency", "user").all()
    return render(request, "cabinet/manage_clients.html", {"form": form, "clients": clients})


@login_required
@require_http_methods(["POST"])
def api_admin_reset_agency_owner(request, agency_id: int):
    # Only superusers
    if not request.user.is_superuser:
        return JsonResponse({"success": False, "error": "Forbidden"}, status=403)
    agency = Agency.objects.filter(id=agency_id).first()
    if not agency:
        return JsonResponse({"success": False, "error": "Agency not found"}, status=404)
    from django.contrib.auth import get_user_model
    from django.utils.text import slugify
    import secrets
    User = get_user_model()
    password = secrets.token_urlsafe(9)
    base_username = slugify(agency.name) or f"agency{agency.id}"
    username = base_username
    suffix = 1
    owner_id = getattr(agency.owner, "id", None)
    while User.objects.filter(username=username).exclude(id=owner_id).exists():
        suffix += 1
        username = f"{base_username}{suffix}"
    # If current owner is a superuser, DO NOT modify it. Create a dedicated non-admin owner.
    if agency.owner and agency.owner.is_superuser:
        user = User.objects.create_user(username=username, password=password)
        agency.owner = user
        agency.save(update_fields=["owner"])
    elif agency.owner:
        user = agency.owner
        if user.username != username:
            user.username = username
        user.set_password(password)
        user.save(update_fields=["username", "password"])
    else:
        user = User.objects.create_user(username=username, password=password)
        agency.owner = user
        agency.save(update_fields=["owner"])
    return JsonResponse({"success": True, "username": username, "password": password})


@login_required
def manage_hashtags(request):
    if not request.user.is_superuser:
        return redirect("/")
    if request.method == "POST":
        form = ClientHashtagForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Hashtag added to client")
            return redirect("cabinet_manage_hashtags")
    else:
        form = ClientHashtagForm()
    items = ClientHashtag.objects.select_related("client", "client__agency").all()
    return render(request, "cabinet/manage_hashtags.html", {"form": form, "items": items})


@login_required
def delete_agency(request, agency_id: int):
    if not request.user.is_superuser:
        return redirect("cabinet_dashboard")
    agency = get_object_or_404(Agency, id=agency_id)
    if request.method == "POST":
        agency.delete()
        messages.success(request, "Agency deleted")
        return redirect("cabinet_manage_agencies")
    return render(request, "cabinet/confirm_delete.html", {"object": agency, "type": "Agency"})


@login_required
def delete_client(request, client_id: int):
    if not request.user.is_superuser:
        return redirect("cabinet_dashboard")
    client = get_object_or_404(Client, id=client_id)
    if request.method == "POST":
        client.delete()
        messages.success(request, "Client deleted")
        return redirect("cabinet_manage_clients")
    return render(request, "cabinet/confirm_delete.html", {"object": client, "type": "Client"})


@login_required
def delete_client_hashtag(request, item_id: int):
    if not request.user.is_superuser:
        return redirect("cabinet_dashboard")
    item = get_object_or_404(ClientHashtag, id=item_id)
    if request.method == "POST":
        item.delete()
        messages.success(request, "Hashtag unassigned from client")
        return redirect("cabinet_manage_hashtags")
    return render(request, "cabinet/confirm_delete.html", {"object": item, "type": "Client Hashtag"})


@login_required
def reset_client_user_password(request, client_id: int):
    if not request.user.is_superuser:
        return redirect("cabinet_dashboard")
    
    client = get_object_or_404(Client, id=client_id)
    
    if request.method == "POST":
        new_password = request.POST.get("new_password", "").strip()
        # If no password provided or too short, generate a safe one
        generate = False
        if not new_password or len(new_password) < 6:
            import secrets
            new_password = secrets.token_urlsafe(9)
            generate = True

        from django.contrib.auth import get_user_model
        User = get_user_model()

        # If no user or current user is superuser, create dedicated non-admin user
        if not client.user or client.user.is_superuser:
            base_username = slugify(client.name or "client") or "client"
            username = base_username
            suffix = 1
            while User.objects.filter(username=username).exists():
                suffix += 1
                username = f"{base_username}{suffix}"
            login_user = User.objects.create_user(username=username, password=new_password)
            client.user = login_user
            client.save(update_fields=["user"])
            messages.success(
                request,
                f"Client login created. Credentials — Login: {username} | Password: {new_password}"
            )
        else:
            user = client.user
            user.set_password(new_password)
            user.save(update_fields=["password"])
            if generate:
                messages.success(request, f"Password reset. New Password: {new_password}")
            else:
                messages.success(request, f"Password reset for {user.username}")
        return redirect("cabinet_manage_clients")
    
    return render(request, "cabinet/reset_password.html", {"client": client})


@login_required
def admin_dashboard(request):
    # Restrict direct access: only superusers
    if not request.user.is_superuser:
        return redirect("cabinet_dashboard")
    agencies = Agency.objects.select_related("owner").all()
    clients = Client.objects.select_related("agency", "user").all()
    # Top hashtags by latest snapshot views
    top_analytics = (
        HashtagAnalytics.objects.order_by("hashtag", "-created_at")
    )
    # Deduplicate to latest per hashtag
    seen = set()
    latest_unique: list[HashtagAnalytics] = []
    for row in top_analytics:
        if row.hashtag not in seen:
            latest_unique.append(row)
            seen.add(row.hashtag)
        if len(latest_unique) >= 20:
            break
    # Defaults for KPIs/analytics to avoid NameError on edge branches
    kpi_total_views = 0
    kpi_total_videos = 0
    kpi_avg_per_video = 0.0
    hashtags_breakdown = []
    clients_breakdown = []
    accounts_analytics = []

    # Daily stats for last 7 days (views and videos)
    end = timezone.now()
    start = end - timezone.timedelta(days=7)
    qs = (
        HashtagAnalytics.objects.filter(created_at__gte=start, created_at__lte=end)
        .annotate(day=TruncDate("created_at"))
        .values("day")
        .annotate(
            total_views=Sum("total_views"),
            total_videos=Sum("analyzed_medias"),
            total_likes=Sum("total_likes"),
            total_comments=Sum("total_comments"),
        )
        .order_by("day")
    )
    daily_stats = [
        {
            "date": row["day"].strftime("%Y-%m-%d") if row["day"] else "",
            "day_name": row["day"].strftime("%b %d") if row["day"] else "",
            "total_views": int(row.get("total_views") or 0),
            "total_videos": int(row.get("total_videos") or 0),
            "total_likes": int(row.get("total_likes") or 0),
            "total_comments": int(row.get("total_comments") or 0),
        }
        for row in qs
    ]

    # KPI: totals and avg per video (last 7 days)
    kpi_total_views = sum(d["total_views"] for d in daily_stats)
    kpi_total_videos = sum(d["total_videos"] for d in daily_stats)
    kpi_avg_per_video = (kpi_total_views / kpi_total_videos) if kpi_total_videos > 0 else 0.0

    # Agencies analytics (last 7 days)
    agencies_rows: list[dict] = []
    for a in agencies:
        a_clients = list(a.clients.all())
        a_hashtags = ClientHashtag.objects.filter(client__in=a_clients).values_list("hashtag", flat=True)
        if not a_hashtags:
            agencies_rows.append({
                "agency": {"id": a.id, "name": a.name},
                "clients_count": len(a_clients),
                "views": 0,
                "videos": 0,
                "avg_views": 0.0,
            })
            continue
        agg = (
            HashtagAnalytics.objects.filter(hashtag__in=list(a_hashtags), created_at__gte=start, created_at__lte=end)
            .aggregate(
                views=Sum("total_views"),
                videos=Sum("analyzed_medias"),
                likes=Sum("total_likes"),
                comments=Sum("total_comments"),
            )
        )
        views_sum = int(agg.get("views") or 0)
        videos_sum = int(agg.get("videos") or 0)
        avg_views = (views_sum / videos_sum) if videos_sum > 0 else 0.0
        agencies_rows.append({
            "agency": {"id": a.id, "name": a.name},
            "clients_count": len(a_clients),
            "views": views_sum,
            "videos": videos_sum,
            "avg_views": avg_views,
        })
    agencies_rows = sorted(agencies_rows, key=lambda x: x["views"], reverse=True)

    # Clients analytics (last 7 days, top 50)
    clients_rows: list[dict] = []
    for c in clients:
        c_hashtags = ClientHashtag.objects.filter(client=c).values_list("hashtag", flat=True)
        if not c_hashtags:
            continue
        agg = (
            HashtagAnalytics.objects.filter(hashtag__in=list(c_hashtags), created_at__gte=start, created_at__lte=end)
            .aggregate(
                views=Sum("total_views"),
                videos=Sum("analyzed_medias"),
                likes=Sum("total_likes"),
                comments=Sum("total_comments"),
            )
        )
        views_sum = int(agg.get("views") or 0)
        videos_sum = int(agg.get("videos") or 0)
        if views_sum == 0 and videos_sum == 0:
            continue
        avg_views = (views_sum / videos_sum) if videos_sum > 0 else 0.0
        clients_rows.append({
            "client": {"id": c.id, "name": c.name},
            "agency": {"id": c.agency.id, "name": c.agency.name} if c.agency else None,
            "views": views_sum,
            "videos": videos_sum,
            "avg_views": avg_views,
        })
    clients_rows = sorted(clients_rows, key=lambda x: x["views"], reverse=True)[:50]

    return render(
        request,
        "cabinet/admin_dashboard.html",
        {
            "agencies": [{"id": a.id, "name": a.name} for a in agencies],
            "clients": [{"id": c.id, "name": c.name} for c in clients],
            "top_hashtags": [
                {
                    "hashtag": h.hashtag,
                    "total_views": h.total_views or 0,
                    "analyzed_medias": h.analyzed_medias or 0,
                    "average_views": h.average_views or 0.0,
                    "total_likes": getattr(h, 'total_likes', 0) or 0,
                    "total_comments": getattr(h, 'total_comments', 0) or 0,
                    "engagement_rate": getattr(h, 'engagement_rate', 0.0) or 0.0,
                    "created_at": h.created_at.isoformat() if h.created_at else "",
                }
                for h in latest_unique
            ],
            "daily_stats": daily_stats,
            "agencies_rows": agencies_rows,
            "clients_rows": clients_rows,
        },
    )


@login_required
def agency_dashboard(request):
    # Defaults for KPIs/analytics to avoid NameError on edge branches
    kpi_total_views = 0
    kpi_total_videos = 0
    kpi_avg_per_video = 0.0
    hashtags_breakdown = []
    clients_breakdown = []
    accounts_analytics = []

    agency = Agency.objects.filter(owner=request.user).first()
    if not agency:
        # Check if user is a client - if so, redirect to personal dashboard
        client_for_scope = Client.objects.filter(user=request.user).select_related("agency").first()
        if client_for_scope:
            return redirect("cabinet_client_dashboard")
        
        return render(request, "cabinet/error.html", {"message": "No agency found for this user."})

    # Build client list (all clients for agency owner)
    clients_list = list(agency.clients.select_related("user").all())

    # Get time period parameter
    days_param = request.GET.get("days")
    if days_param:
        try:
            days = int(days_param)
        except ValueError:
            days = None
    else:
        days = None  # All time by default

    client_rows = []
    for c in clients_list:
        # Use AnalyticsService to get combined data
        analytics_service = AnalyticsService(c)
        combined_summary = analytics_service.get_combined_analytics_summary(days=days)
        network_breakdown = analytics_service.get_network_breakdown(days=days)
        
        # Calculate total accounts across all networks
        total_accounts = sum(network.total_accounts for network in combined_summary.networks.values())
        
        client_rows.append({
            "client": c, 
            "total_views": combined_summary.total_views, 
            "total_videos": combined_summary.total_posts, 
            "avg_views": combined_summary.average_views,
            "engagement_rate": combined_summary.engagement_rate,
            "total_accounts": total_accounts,
            "network_breakdown": network_breakdown,
            "networks_count": len(network_breakdown)
        })

    # Daily stats for last 7 days (views and videos)
    end = timezone.now()
    start = end - timezone.timedelta(days=7)
    agency_hashtags = ClientHashtag.objects.filter(client__agency=agency).values_list("hashtag", flat=True)
    # Daily stats aggregated by day
    qs = (
        HashtagAnalytics.objects.filter(hashtag__in=agency_hashtags, created_at__gte=start, created_at__lte=end)
        .annotate(day=TruncDate("created_at"))
        .values("day")
        .annotate(
            total_views=Sum("total_views"),
            total_videos=Sum("analyzed_medias"),
            total_likes=Sum("total_likes"),
            total_comments=Sum("total_comments"),
        )
        .order_by("day")
    )
    daily_stats = [
        {
            "date": row["day"].strftime("%Y-%m-%d") if row["day"] else "",
            "day_name": row["day"].strftime("%b %d") if row["day"] else "",
            "total_views": int(row.get("total_views") or 0),
            "total_videos": int(row.get("total_videos") or 0),
            "total_likes": int(row.get("total_likes") or 0),
            "total_comments": int(row.get("total_comments") or 0),
        }
        for row in qs
    ]

    # Daily stats with hashtag breakdown for tooltips
    qs_detailed = (
        HashtagAnalytics.objects.filter(hashtag__in=agency_hashtags, created_at__gte=start, created_at__lte=end)
        .annotate(day=TruncDate("created_at"))
        .values("day", "hashtag")
        .annotate(
            views=Sum("total_views"),
            videos=Sum("analyzed_medias"),
            likes=Sum("total_likes"),
            comments=Sum("total_comments"),
        )
        .order_by("day", "hashtag")
    )
    
    # Group by day with hashtag details for tooltips
    daily_stats_detailed = {}
    for row in qs_detailed:
        day_str = row["day"].strftime("%Y-%m-%d") if row["day"] else ""
        if day_str not in daily_stats_detailed:
            daily_stats_detailed[day_str] = {
                "date": day_str,
                "day_name": row["day"].strftime("%b %d") if row["day"] else "",
                "hashtags": []
            }
        daily_stats_detailed[day_str]["hashtags"].append({
            "hashtag": row["hashtag"],
            "views": int(row.get("views") or 0),
            "videos": int(row.get("videos") or 0),
            "likes": int(row.get("likes") or 0),
            "comments": int(row.get("comments") or 0),
        })

    # KPIs - sum from client analytics
    kpi_total_views = sum(c["total_views"] for c in client_rows)
    kpi_total_videos = sum(c["total_videos"] for c in client_rows)
    kpi_avg_per_video = (kpi_total_views / kpi_total_videos) if kpi_total_videos > 0 else 0.0

    # Hashtag breakdown for last 7 days (top 10 by views)
    hb_qs = (
        HashtagAnalytics.objects.filter(hashtag__in=agency_hashtags, created_at__gte=start, created_at__lte=end)
        .values("hashtag")
        .annotate(
            views=Sum("total_views"),
            videos=Sum("analyzed_medias"),
            likes=Sum("total_likes"),
            comments=Sum("total_comments"),
        )
        .order_by("-views")[:10]
    )
    hashtags_breakdown = [
        {
            "hashtag": row["hashtag"],
            "views": int(row.get("views") or 0),
            "videos": int(row.get("videos") or 0),
            "likes": int(row.get("likes") or 0),
            "comments": int(row.get("comments") or 0),
            "avg_views": (int(row.get("views") or 0) / int(row.get("videos") or 0)) if int(row.get("videos") or 0) > 0 else 0.0,
            "er": ((int(row.get("likes") or 0) + int(row.get("comments") or 0)) / max(int(row.get("views") or 0), 1)),
        }
        for row in hb_qs
    ]

    # Client breakdown by latest snapshot totals (sum of last snapshot per client's hashtags)
    clients_breakdown = []
    for c in clients_list:
        c_hashtags = ClientHashtag.objects.filter(client=c).values_list("hashtag", flat=True)
        latest_views = 0
        latest_videos = 0
        for h in c_hashtags:
            last = HashtagAnalytics.objects.filter(hashtag=h).order_by("-created_at").first()
            if last:
                latest_views += int(last.total_views or 0)
                latest_videos += int(getattr(last, "analyzed_medias", 0) or 0)
        clients_breakdown.append({
            "client": c.name,
            "views": latest_views,
            "videos": latest_videos,
        })
    # Account analytics (last 7 days) for accounts under agency's clients
    accounts_qs = (
        AccountAnalytics.objects.filter(
            account__client__in=clients_list, created_at__gte=start, created_at__lte=end
        )
        .values("account__id", "account__username")
        .annotate(
            views=Sum("total_views"),
            videos=Sum("total_videos"),
        )
        .order_by("-views")[:15]
    )
    accounts_analytics = [
        {
            "account_id": row["account__id"],
            "username": row["account__username"] or f"#{row['account__id']}",
            "views": int(row.get("views") or 0),
            "videos": int(row.get("videos") or 0),
            "avg_views": (int(row.get("views") or 0) / int(row.get("videos") or 0)) if int(row.get("videos") or 0) > 0 else 0.0,
        }
        for row in accounts_qs
    ]
    # Sort and keep top 10 by views
    clients_breakdown = sorted(clients_breakdown, key=lambda x: x["views"], reverse=True)[:10]

    # Get recent calculations for this agency
    recent_calculations = CalculationHistory.objects.filter(
        agency=agency
    ).select_related('user', 'agency').order_by('-created_at')[:10]

    return render(
        request,
        "cabinet/agency_dashboard.html",
        {
            "agency": agency,
            "client_rows": client_rows,
            "daily_stats": daily_stats,
            "daily_stats_detailed": daily_stats_detailed,
            "hashtags_breakdown": hashtags_breakdown,
            "clients_breakdown": clients_breakdown,
            "kpi_total_views": kpi_total_views,
            "kpi_total_videos": kpi_total_videos,
            "kpi_avg_per_video": kpi_avg_per_video,
            "accounts_analytics": accounts_analytics,
            "recent_calculations": recent_calculations,
        },
    )


def export_excel(request):
    """Export detailed analytics to Excel with multiple sheets and comprehensive data."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.db.models import Sum, Avg, Max, Count
    from datetime import datetime, timedelta
    
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)

    user = request.user
    filename = f"detailed_analytics_{user.username}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    if user.is_superuser:
        # Sheet 1: Summary Overview
        ws_summary = wb.create_sheet("Summary Overview")
        ws_summary.append(["Analytics Summary", f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"])
        ws_summary.append([])
        
        # Get total counts
        total_clients = Client.objects.count()
        total_agencies = Agency.objects.count()
        total_hashtags = ClientHashtag.objects.count()
        total_analytics = HashtagAnalytics.objects.count()
        manual_analytics = HashtagAnalytics.objects.filter(is_manual=True).count()
        auto_analytics = HashtagAnalytics.objects.filter(is_manual=False).count()
        
        ws_summary.append(["Total Clients", total_clients])
        ws_summary.append(["Total Agencies", total_agencies])
        ws_summary.append(["Total Hashtags", total_hashtags])
        ws_summary.append(["Total Analytics Records", total_analytics])
        
        # Sheet 2: Clients Overview
        ws_clients = wb.create_sheet("Clients Overview")
        headers = ["Client Name", "Agency", "Total Hashtags", "Total Analytics", "Last Activity"]
        ws_clients.append(headers)
        
        # Style headers
        for col in range(1, len(headers) + 1):
            cell = ws_clients.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        clients = Client.objects.select_related('agency').prefetch_related('client_hashtags', 'hashtag_analytics')
        for client in clients:
            hashtag_count = client.client_hashtags.count()
            analytics_count = client.hashtag_analytics.count()
            last_activity = client.hashtag_analytics.order_by('-created_at').first()
            last_activity_date = last_activity.created_at.strftime('%Y-%m-%d') if last_activity else 'No activity'
            
            ws_clients.append([
                client.name,
                client.agency.name if client.agency else 'No agency',
                hashtag_count,
                analytics_count,
                last_activity_date
            ])
        
        # Sheet 3: Analytics Details
        ws_manual = wb.create_sheet("Analytics Details")
        headers = ["Client", "Social Network", "Hashtag", "Posts", "Views", "Likes", "Comments", "Shares", 
                  "Followers", "Growth Rate", "Accounts", "Avg Views/Video", "Avg Likes/Video", "Created At"]
        ws_manual.append(headers)
        
        # Style headers
        for col in range(1, len(headers) + 1):
            cell = ws_manual.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        manual_analytics = HashtagAnalytics.objects.select_related('client').order_by('-created_at')
        for analytics in manual_analytics:
            ws_manual.append([
                analytics.client.name if analytics.client else 'No client',
                analytics.get_social_network_display(),
                f"#{analytics.hashtag}" if analytics.hashtag else 'No hashtag',
                analytics.analyzed_medias or 0,
                analytics.total_views or 0,
                analytics.total_likes or 0,
                analytics.total_comments or 0,
                analytics.total_shares or 0,
                analytics.total_followers or 0,
                f"{(analytics.growth_rate or 0):.2f}%" if analytics.growth_rate else '0.00%',
                analytics.total_accounts or 0,
                f"{(analytics.avg_views_per_video or 0):.1f}" if analytics.avg_views_per_video else '0.0',
                f"{(analytics.avg_likes_per_video or 0):.1f}" if analytics.avg_likes_per_video else '0.0',
                analytics.created_at.strftime('%Y-%m-%d %H:%M')
            ])
        
        # Sheet 4: Social Network Breakdown
        ws_networks = wb.create_sheet("Social Networks")
        headers = ["Social Network", "Total Records", "Total Posts", "Total Views", "Total Likes", 
                  "Avg Views/Post", "Avg Likes/Post", "Total Accounts", "Avg Posts/Account"]
        ws_networks.append(headers)
        
        # Style headers
        for col in range(1, len(headers) + 1):
            cell = ws_networks.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Aggregate by social network
        network_stats = HashtagAnalytics.objects.values('social_network').annotate(
            total_records=Count('id'),
            total_posts=Sum('analyzed_medias'),
            total_views=Sum('total_views'),
            total_likes=Sum('total_likes'),
            total_accounts=Sum('total_accounts'),
            avg_views_per_post=Avg('avg_views_per_video'),
            avg_likes_per_post=Avg('avg_likes_per_video'),
            avg_posts_per_account=Avg('avg_videos_per_account')
        ).order_by('social_network')
        
        for stat in network_stats:
            ws_networks.append([
                stat['social_network'],
                stat['total_records'] or 0,
                stat['total_posts'] or 0,
                stat['total_views'] or 0,
                stat['total_likes'] or 0,
                f"{(stat['avg_views_per_post'] or 0):.1f}",
                f"{(stat['avg_likes_per_post'] or 0):.1f}",
                stat['total_accounts'] or 0,
                f"{(stat['avg_posts_per_account'] or 0):.1f}"
            ])
        
        # Sheet 5: Daily Analytics (Last 30 days)
        ws_daily = wb.create_sheet("Daily Analytics")
        headers = ["Date", "Client", "Social Network", "Hashtag", "Posts", "Views", "Likes", "Comments", "Shares", "Followers", "Accounts"]
        ws_daily.append(headers)
        
        # Style headers
        for col in range(1, len(headers) + 1):
            cell = ws_daily.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        thirty_days_ago = timezone.now() - timedelta(days=30)
        daily_analytics = HashtagAnalytics.objects.filter(
            created_at__gte=thirty_days_ago
        ).select_related('client').order_by('-created_at')
        
        for analytics in daily_analytics:
            ws_daily.append([
                analytics.created_at.strftime('%Y-%m-%d'),
                analytics.client.name if analytics.client else 'No client',
                analytics.get_social_network_display(),
                f"#{analytics.hashtag}" if analytics.hashtag else 'No hashtag',
                analytics.analyzed_medias or 0,
                analytics.total_views or 0,
                analytics.total_likes or 0,
                analytics.total_comments or 0,
                analytics.total_shares or 0,
                analytics.total_followers or 0,
                analytics.total_accounts or 0
            ])
        
        # Sheet 6: Daily Aggregation by Date
        ws_daily_agg = wb.create_sheet("Daily Aggregation")
        headers = ["Date", "Total Posts", "Total Views", "Total Likes", "Total Comments", "Total Shares", "Total Accounts", "Records Count"]
        ws_daily_agg.append(headers)
        
        # Style headers
        for col in range(1, len(headers) + 1):
            cell = ws_daily_agg.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Aggregate by date
        from django.db.models import Sum, Count
        daily_stats = HashtagAnalytics.objects.filter(
            created_at__gte=thirty_days_ago
        ).extra(select={'date': 'DATE(created_at)'}).values('date').annotate(
            total_posts=Sum('analyzed_medias'),
            total_views=Sum('total_views'),
            total_likes=Sum('total_likes'),
            total_comments=Sum('total_comments'),
            total_shares=Sum('total_shares'),
            total_accounts=Sum('total_accounts'),
            records_count=Count('id')
        ).order_by('-date')
        
        for stat in daily_stats:
            ws_daily_agg.append([
                stat['date'].strftime('%Y-%m-%d'),
                stat['total_posts'] or 0,
                stat['total_views'] or 0,
                stat['total_likes'] or 0,
                stat['total_comments'] or 0,
                stat['total_shares'] or 0,
                stat['total_accounts'] or 0,
                stat['records_count']
            ])
        
        # Sheet 7: Client Performance Comparison
        ws_comparison = wb.create_sheet("Client Comparison")
        headers = ["Client", "Agency", "Total Posts", "Total Views", "Avg Views/Post", "Total Likes", "Total Accounts", "Networks Count", "Last Activity"]
        ws_comparison.append(headers)
        
        # Style headers
        for col in range(1, len(headers) + 1):
            cell = ws_comparison.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Get all clients with their analytics
        all_clients = Client.objects.select_related('agency').prefetch_related('hashtag_analytics')
        for client in all_clients:
            # Get client's analytics for last 30 days
            client_analytics = client.hashtag_analytics.filter(
                created_at__gte=thirty_days_ago
            )
            
            total_posts = sum(a.analyzed_medias or 0 for a in client_analytics)
            total_views = sum(a.total_views or 0 for a in client_analytics)
            total_likes = sum(a.total_likes or 0 for a in client_analytics)
            total_accounts = sum(a.total_accounts or 0 for a in client_analytics)
            
            # Get unique networks
            networks = set(a.social_network for a in client_analytics if a.social_network)
            
            # Get last activity
            last_activity = client_analytics.order_by('-created_at').first()
            last_activity_date = last_activity.created_at.strftime('%Y-%m-%d') if last_activity else 'No activity'
            
            avg_views = (total_views / total_posts) if total_posts > 0 else 0
            
            ws_comparison.append([
                client.name,
                client.agency.name if client.agency else 'No agency',
                total_posts,
                total_views,
                f"{avg_views:.1f}",
                total_likes,
                total_accounts,
                len(networks),
                last_activity_date
            ])
        
        # Sheet 8: Hashtag Performance
        ws_hashtags = wb.create_sheet("Hashtag Performance")
        headers = ["Hashtag", "Client", "Social Network", "Total Posts", "Total Views", "Avg Views/Post", "Total Likes", "Total Accounts", "Last Update"]
        ws_hashtags.append(headers)
        
        # Style headers
        for col in range(1, len(headers) + 1):
            cell = ws_hashtags.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Get hashtag performance
        hashtag_stats = HashtagAnalytics.objects.filter(
            created_at__gte=thirty_days_ago
        ).select_related('client').values('hashtag', 'client__name', 'social_network').annotate(
            total_posts=Sum('analyzed_medias'),
            total_views=Sum('total_views'),
            total_likes=Sum('total_likes'),
            total_accounts=Sum('total_accounts'),
            last_update=Max('created_at')
        ).order_by('-total_views')
        
        for stat in hashtag_stats:
            avg_views = (stat['total_views'] / stat['total_posts']) if stat['total_posts'] > 0 else 0
            ws_hashtags.append([
                f"#{stat['hashtag']}" if stat['hashtag'] else 'No hashtag',
                stat['client__name'] or 'No client',
                dict(HashtagAnalytics.SOCIAL_NETWORK_CHOICES).get(stat['social_network'], stat['social_network']),
                stat['total_posts'] or 0,
                stat['total_views'] or 0,
                f"{avg_views:.1f}",
                stat['total_likes'] or 0,
                stat['total_accounts'] or 0,
                stat['last_update'].strftime('%Y-%m-%d') if stat['last_update'] else 'No date'
            ])
        
        # Sheet 9: Platform-Specific Metrics
        ws_platform = wb.create_sheet("Platform Metrics")
        headers = ["Social Network", "Metric", "Value", "Description"]
        ws_platform.append(headers)
        
        # Style headers
        for col in range(1, len(headers) + 1):
            cell = ws_platform.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Platform-specific metrics
        platform_metrics = [
            ('INSTAGRAM', 'Stories Views', 'instagram_stories_views', 'Total Instagram Stories views'),
            ('INSTAGRAM', 'Reels Views', 'instagram_reels_views', 'Total Instagram Reels views'),
            ('YOUTUBE', 'Subscribers', 'youtube_subscribers', 'Total YouTube subscribers'),
            ('YOUTUBE', 'Watch Time (min)', 'youtube_watch_time', 'Total YouTube watch time in minutes'),
            ('TIKTOK', 'Video Views', 'tiktok_video_views', 'Total TikTok video views'),
            ('TIKTOK', 'Profile Views', 'tiktok_profile_views', 'Total TikTok profile views'),
        ]
        
        for network, metric_name, field_name, description in platform_metrics:
            total_value = HashtagAnalytics.objects.filter(
                social_network=network,
                created_at__gte=thirty_days_ago
            ).aggregate(total=Sum(field_name))['total'] or 0
            
            ws_platform.append([
                dict(HashtagAnalytics.SOCIAL_NETWORK_CHOICES).get(network, network),
                metric_name,
                total_value,
                description
            ])
    
    else:
        # For non-superusers, show only their data
        client = Client.objects.filter(user=user).first()
        if client:
            service = AnalyticsService(client)
            
            # Sheet 1: Client Summary
            ws_summary = wb.create_sheet("My Analytics")
            ws_summary.append([f"Analytics for {client.name}", f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"])
            ws_summary.append([])
            
            # Get client's analytics summary
            networks = service.get_manual_analytics_by_network(30)
            total_posts = sum(network.total_posts for network in networks.values())
            total_views = sum(network.total_views for network in networks.values())
            total_likes = sum(network.total_likes for network in networks.values())
            
            ws_summary.append(["Total Posts (30 days)", total_posts])
            ws_summary.append(["Total Views (30 days)", total_views])
            ws_summary.append(["Total Likes (30 days)", total_likes])
            ws_summary.append(["Networks Active", len(networks)])
            
            # Sheet 2: Network Breakdown
            ws_networks = wb.create_sheet("By Network")
            headers = ["Network", "Posts", "Views", "Likes", "Avg Views", "Engagement", "Accounts"]
            ws_networks.append(headers)
            
            for network_key, network_data in networks.items():
                ws_networks.append([
                    network_data.network,
                    network_data.total_posts,
                    network_data.total_views,
                    network_data.total_likes,
                    f"{network_data.average_views:.1f}",
                    f"{network_data.engagement_rate:.2%}",
                    network_data.total_accounts
                ])
            
            # Sheet 3: Detailed Records (only if data exists)
            detailed_records = HashtagAnalytics.objects.filter(
                client=client,
                created_at__gte=timezone.now() - timedelta(days=30)
            ).order_by('-created_at')
            
            if detailed_records.exists():
                ws_details = wb.create_sheet("Detailed Records")
                headers = ["Date", "Network", "Hashtag", "Posts", "Views", "Likes", "Comments", "Shares", "Followers", "Accounts"]
                ws_details.append(headers)
                
                for record in detailed_records:
                    ws_details.append([
                        record.created_at.strftime('%Y-%m-%d'),
                        record.get_social_network_display(),
                        f"#{record.hashtag}" if record.hashtag else 'No hashtag',
                        record.analyzed_medias or 0,
                        record.total_views or 0,
                        record.total_likes or 0,
                        record.total_comments or 0,
                        record.total_shares or 0,
                        record.total_followers or 0,
                        record.total_accounts or 0
                    ])
            
            # Sheet 4: Daily Performance (only if data exists)
            from django.db.models import Sum, Count
            daily_stats = HashtagAnalytics.objects.filter(
                client=client,
                created_at__gte=timezone.now() - timedelta(days=30)
            ).extra(select={'date': 'DATE(created_at)'}).values('date').annotate(
                total_posts=Sum('analyzed_medias'),
                total_views=Sum('total_views'),
                total_likes=Sum('total_likes'),
                total_comments=Sum('total_comments'),
                total_shares=Sum('total_shares'),
                total_accounts=Sum('total_accounts'),
                records_count=Count('id')
            ).order_by('-date')
            
            if daily_stats.exists():
                ws_daily = wb.create_sheet("Daily Performance")
                headers = ["Date", "Posts", "Views", "Likes", "Comments", "Shares", "Accounts", "Records"]
                ws_daily.append(headers)
                
                for stat in daily_stats:
                    ws_daily.append([
                        stat['date'].strftime('%Y-%m-%d'),
                        stat['total_posts'] or 0,
                        stat['total_views'] or 0,
                        stat['total_likes'] or 0,
                        stat['total_comments'] or 0,
                        stat['total_shares'] or 0,
                        stat['total_accounts'] or 0,
                        stat['records_count']
                    ])
            
            # Sheet 5: Hashtag Breakdown (only if data exists)
            hashtag_stats = HashtagAnalytics.objects.filter(
                client=client,
                created_at__gte=timezone.now() - timedelta(days=30)
            ).values('hashtag', 'social_network').annotate(
                total_posts=Sum('analyzed_medias'),
                total_views=Sum('total_views'),
                total_likes=Sum('total_likes'),
                total_accounts=Sum('total_accounts'),
                last_update=Max('created_at')
            ).order_by('-total_views')
            
            if hashtag_stats.exists():
                ws_hashtags = wb.create_sheet("Hashtag Breakdown")
                headers = ["Hashtag", "Network", "Posts", "Views", "Avg Views", "Likes", "Accounts", "Last Update"]
                ws_hashtags.append(headers)
                
                for stat in hashtag_stats:
                    avg_views = (stat['total_views'] / stat['total_posts']) if stat['total_posts'] > 0 else 0
                    ws_hashtags.append([
                        f"#{stat['hashtag']}" if stat['hashtag'] else 'No hashtag',
                        dict(HashtagAnalytics.SOCIAL_NETWORK_CHOICES).get(stat['social_network'], stat['social_network']),
                        stat['total_posts'] or 0,
                        stat['total_views'] or 0,
                        f"{avg_views:.1f}",
                        stat['total_likes'] or 0,
                        stat['total_accounts'] or 0,
                        stat['last_update'].strftime('%Y-%m-%d') if stat['last_update'] else 'No date'
                    ])
            
            # Sheet 6: Advanced Metrics (only if data exists)
            advanced_records = HashtagAnalytics.objects.filter(
                client=client,
                created_at__gte=timezone.now() - timedelta(days=30)
            ).order_by('-created_at')
            
            if advanced_records.exists():
                ws_advanced = wb.create_sheet("Advanced Metrics")
                headers = ["Network", "Hashtag", "Avg Videos/Account", "Max Videos/Account", "Avg Views/Video", "Max Views/Video", 
                          "Avg Views/Account", "Max Views/Account", "Avg Likes/Video", "Max Likes/Video", "Avg Likes/Account", "Max Likes/Account"]
                ws_advanced.append(headers)
                
                for record in advanced_records:
                    ws_advanced.append([
                        record.get_social_network_display(),
                        f"#{record.hashtag}" if record.hashtag else 'No hashtag',
                        f"{(record.avg_videos_per_account or 0):.1f}",
                        record.max_videos_per_account or 0,
                        f"{(record.avg_views_per_video or 0):.1f}",
                        record.max_views_per_video or 0,
                        f"{(record.avg_views_per_account or 0):.1f}",
                        record.max_views_per_account or 0,
                        f"{(record.avg_likes_per_video or 0):.1f}",
                        record.max_likes_per_video or 0,
                        f"{(record.avg_likes_per_account or 0):.1f}",
                        record.max_likes_per_account or 0
                    ])
        
        else:
            agency = Agency.objects.filter(owner=user).first()
            if agency:
                # Sheet 1: Agency Summary
                ws_summary = wb.create_sheet("Agency Summary")
                ws_summary.append([f"Agency: {agency.name}", f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"])
                ws_summary.append([])
                
                clients = agency.clients.all()
                total_clients = clients.count()
                
                # Calculate agency totals
                total_views = 0
                total_posts = 0
                total_accounts = 0
                
                for c in clients:
                    analytics_service = AnalyticsService(c)
                    combined_summary = analytics_service.get_combined_analytics_summary(days=30)
                    total_views += combined_summary.total_views
                    total_posts += combined_summary.total_posts
                    total_accounts += sum(network.total_accounts for network in combined_summary.networks.values())
                
                ws_summary.append(["Total Clients", total_clients])
                ws_summary.append(["Total Views (30 days)", total_views])
                ws_summary.append(["Total Posts (30 days)", total_posts])
                ws_summary.append(["Total Accounts", total_accounts])
                
                # Sheet 2: Client Comparison
                ws_clients = wb.create_sheet("Client Comparison")
                headers = ["Client", "Total Views", "Total Posts", "Avg Views/Post", "Total Accounts", "Networks", "Last Activity"]
                ws_clients.append(headers)
                
                for c in clients:
                    analytics_service = AnalyticsService(c)
                    combined_summary = analytics_service.get_combined_analytics_summary(days=30)
                    total_accounts = sum(network.total_accounts for network in combined_summary.networks.values())
                    
                    # Get last activity
                    last_activity = c.hashtag_analytics.order_by('-created_at').first()
                    last_activity_date = last_activity.created_at.strftime('%Y-%m-%d') if last_activity else 'No activity'
                    
                    ws_clients.append([
                        c.name,
                        combined_summary.total_views,
                        combined_summary.total_posts,
                        f"{combined_summary.average_views:.1f}",
                        total_accounts,
                        len(combined_summary.networks),
                        last_activity_date
                    ])
                
                # Sheet 3: Network Breakdown (by all clients)
                ws_networks = wb.create_sheet("By Network")
                headers = ["Network", "Posts", "Views", "Likes", "Comments", "Shares", "Followers", "Avg Views", "Engagement", "Accounts"]
                ws_networks.append(headers)
                
                # Aggregate networks across all agency clients
                agency_networks = {}
                for c in clients:
                    analytics_service = AnalyticsService(c)
                    client_networks = analytics_service.get_manual_analytics_by_network(30)
                    for network_key, network_data in client_networks.items():
                        if network_key not in agency_networks:
                            agency_networks[network_key] = {
                                'network': network_data.network,
                                'total_posts': 0,
                                'total_views': 0,
                                'total_likes': 0,
                                'total_comments': 0,
                                'total_shares': 0,
                                'total_followers': 0,
                                'total_accounts': 0
                            }
                        agency_networks[network_key]['total_posts'] += network_data.total_posts
                        agency_networks[network_key]['total_views'] += network_data.total_views
                        agency_networks[network_key]['total_likes'] += network_data.total_likes
                        agency_networks[network_key]['total_comments'] += network_data.total_comments
                        agency_networks[network_key]['total_shares'] += network_data.total_shares
                        agency_networks[network_key]['total_followers'] += network_data.total_followers
                        agency_networks[network_key]['total_accounts'] += network_data.total_accounts
                
                for network_key, network_data in agency_networks.items():
                    avg_views = (network_data['total_views'] / network_data['total_posts']) if network_data['total_posts'] > 0 else 0
                    total_engagement = network_data['total_likes'] + network_data['total_comments'] + network_data['total_shares']
                    engagement_rate = (total_engagement / network_data['total_views']) if network_data['total_views'] > 0 else 0
                    
                    ws_networks.append([
                        network_data['network'],
                        network_data['total_posts'],
                        network_data['total_views'],
                        network_data['total_likes'],
                        network_data['total_comments'],
                        network_data['total_shares'],
                        network_data['total_followers'],
                        f"{avg_views:.1f}",
                        f"{engagement_rate:.2%}",
                        network_data['total_accounts']
                    ])
                
                # Sheet 4: All Client Analytics
                ws_analytics = wb.create_sheet("All Client Analytics")
                headers = ["Client", "Social Network", "Hashtag", "Posts", "Views", "Likes", "Comments", "Shares", "Followers", "Accounts", "Date"]
                ws_analytics.append(headers)
                
                for c in clients:
                    client_analytics = c.hashtag_analytics.filter(
                        created_at__gte=timezone.now() - timedelta(days=30)
                    ).order_by('-created_at')
                    
                    for record in client_analytics:
                        ws_analytics.append([
                            c.name,
                            record.get_social_network_display(),
                            f"#{record.hashtag}" if record.hashtag else 'No hashtag',
                            record.analyzed_medias or 0,
                            record.total_views or 0,
                            record.total_likes or 0,
                            record.total_comments or 0,
                            record.total_shares or 0,
                            record.total_followers or 0,
                            record.total_accounts or 0,
                            record.created_at.strftime('%Y-%m-%d')
                        ])
                
                # Sheet 5: Daily Performance (aggregated across all clients)
                ws_daily = wb.create_sheet("Daily Performance")
                headers = ["Date", "Posts", "Views", "Likes", "Comments", "Shares", "Accounts", "Records"]
                ws_daily.append(headers)
                
                # Aggregate by date for all agency clients
                daily_stats = HashtagAnalytics.objects.filter(
                    client__agency=agency,
                    created_at__gte=timezone.now() - timedelta(days=30)
                ).extra(select={'date': 'DATE(created_at)'}).values('date').annotate(
                    total_posts=Sum('analyzed_medias'),
                    total_views=Sum('total_views'),
                    total_likes=Sum('total_likes'),
                    total_comments=Sum('total_comments'),
                    total_shares=Sum('total_shares'),
                    total_accounts=Sum('total_accounts'),
                    records_count=Count('id')
                ).order_by('-date')
                
                for stat in daily_stats:
                    ws_daily.append([
                        stat['date'].strftime('%Y-%m-%d'),
                        stat['total_posts'] or 0,
                        stat['total_views'] or 0,
                        stat['total_likes'] or 0,
                        stat['total_comments'] or 0,
                        stat['total_shares'] or 0,
                        stat['total_accounts'] or 0,
                        stat['records_count']
                    ])
                
                # Sheet 6: Hashtag Breakdown (across all clients)
                ws_hashtags = wb.create_sheet("Hashtag Breakdown")
                headers = ["Hashtag", "Client", "Network", "Posts", "Views", "Avg Views", "Likes", "Accounts", "Last Update"]
                ws_hashtags.append(headers)
                
                # Get hashtag performance for all agency clients
                hashtag_stats = HashtagAnalytics.objects.filter(
                    client__agency=agency,
                    created_at__gte=timezone.now() - timedelta(days=30)
                ).select_related('client').values('hashtag', 'client__name', 'social_network').annotate(
                    total_posts=Sum('analyzed_medias'),
                    total_views=Sum('total_views'),
                    total_likes=Sum('total_likes'),
                    total_accounts=Sum('total_accounts'),
                    last_update=Max('created_at')
                ).order_by('-total_views')
                
                for stat in hashtag_stats:
                    avg_views = (stat['total_views'] / stat['total_posts']) if stat['total_posts'] > 0 else 0
                    ws_hashtags.append([
                        f"#{stat['hashtag']}" if stat['hashtag'] else 'No hashtag',
                        stat['client__name'] or 'No client',
                        dict(HashtagAnalytics.SOCIAL_NETWORK_CHOICES).get(stat['social_network'], stat['social_network']),
                        stat['total_posts'] or 0,
                        stat['total_views'] or 0,
                        f"{avg_views:.1f}",
                        stat['total_likes'] or 0,
                        stat['total_accounts'] or 0,
                        stat['last_update'].strftime('%Y-%m-%d') if stat['last_update'] else 'No date'
                    ])
                
                # Sheet 7: Advanced Metrics (across all clients)
                ws_advanced = wb.create_sheet("Advanced Metrics")
                headers = ["Client", "Network", "Hashtag", "Avg Videos/Account", "Max Videos/Account", "Avg Views/Video", "Max Views/Video", 
                          "Avg Views/Account", "Max Views/Account", "Avg Likes/Video", "Max Likes/Video", "Avg Likes/Account", "Max Likes/Account"]
                ws_advanced.append(headers)
                
                # Get advanced metrics for all agency clients
                advanced_records = HashtagAnalytics.objects.filter(
                    client__agency=agency,
                    created_at__gte=timezone.now() - timedelta(days=30)
                ).select_related('client').order_by('-created_at')
                
                for record in advanced_records:
                    ws_advanced.append([
                        record.client.name if record.client else 'No client',
                        record.get_social_network_display(),
                        f"#{record.hashtag}" if record.hashtag else 'No hashtag',
                        f"{(record.avg_videos_per_account or 0):.1f}",
                        record.max_videos_per_account or 0,
                        f"{(record.avg_views_per_video or 0):.1f}",
                        record.max_views_per_video or 0,
                        f"{(record.avg_views_per_account or 0):.1f}",
                        record.max_views_per_account or 0,
                        f"{(record.avg_likes_per_video or 0):.1f}",
                        record.max_likes_per_video or 0,
                        f"{(record.avg_likes_per_account or 0):.1f}",
                        record.max_likes_per_account or 0
                    ])
            else:
                # No role assigned - create empty sheet
                ws_summary = wb.create_sheet("No Data")
                ws_summary.append(["No cabinet role assigned", f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"])
                ws_summary.append([])
                ws_summary.append(["Please contact administrator to assign a role."])
    
    # Auto-adjust column widths
    for sheet in wb.worksheets:
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width

    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f"attachment; filename={filename}"
    return resp




@login_required
def search_view(request):
    query = request.GET.get("q", "").strip()
    if request.user.is_superuser:
        agencies = Agency.objects.filter(name__icontains=query) if query else Agency.objects.none()
        clients = Client.objects.filter(name__icontains=query) if query else Client.objects.none()
        hashtag_qs = ClientHashtag.objects.filter(hashtag__icontains=query) if query else ClientHashtag.objects.none()
    else:
        # Scope by role
        my_client = Client.objects.filter(user=request.user).first()
        if my_client:
            agencies = Agency.objects.none()
            clients = Client.objects.none()
            hashtag_qs = ClientHashtag.objects.filter(client=my_client, hashtag__icontains=query) if query else ClientHashtag.objects.none()
        else:
            my_agency = Agency.objects.filter(owner=request.user).first()
            if my_agency:
                agencies = Agency.objects.filter(id=my_agency.id, name__icontains=query) if query else Agency.objects.filter(id=my_agency.id)
                clients = my_agency.clients.filter(name__icontains=query) if query else my_agency.clients.all()
                hashtag_qs = ClientHashtag.objects.filter(client__agency=my_agency, hashtag__icontains=query) if query else ClientHashtag.objects.filter(client__agency=my_agency)
            else:
                agencies = Agency.objects.none()
                clients = Client.objects.none()
                hashtag_qs = ClientHashtag.objects.none()
    hashtags = sorted({ch.hashtag for ch in hashtag_qs})
    return JsonResponse({
        "query": query,
        "results": {
            "agencies": [{"id": a.id, "name": a.name} for a in agencies],
            "clients": [{"id": c.id, "name": c.name} for c in clients],
            "hashtags": [{"name": h} for h in hashtags],
        },
    })


@login_required
def client_data_view(request):
    client_id = request.GET.get("client_id")
    # Resolve client under role restrictions
    if request.user.is_superuser:
        if not client_id:
            return JsonResponse({"error": "Client ID required"}, status=400)
        client = Client.objects.filter(id=client_id).first()
    else:
        my_client = Client.objects.filter(user=request.user).first()
        if my_client:
            client = my_client
        else:
            my_agency = Agency.objects.filter(owner=request.user).first()
            if not my_agency:
                return JsonResponse({"error": "No cabinet role"}, status=403)
            if not client_id:
                return JsonResponse({"error": "Client ID required"}, status=400)
            client = Client.objects.filter(id=client_id, agency=my_agency).first()
    if not client:
        return JsonResponse({"error": "Client not found or not allowed"}, status=404)
    hashtags = list(ClientHashtag.objects.filter(client=client).values_list("hashtag", flat=True))
    # Aggregate by day for last 7 days
    end = timezone.now()
    start = end - timezone.timedelta(days=7)
    qs = (
        HashtagAnalytics.objects.filter(hashtag__in=hashtags, created_at__gte=start, created_at__lte=end)
        .annotate(day=TruncDate("created_at"))
        .values("day")
        .annotate(total_views=Sum("total_views"))
        .order_by("day")
    )
    daily_stats = [
        {
            "date": row["day"].strftime("%Y-%m-%d") if row["day"] else "",
            "day_name": row["day"].strftime("%b %d") if row["day"] else "",
            "total_views": int(row["total_views"] or 0),
        }
        for row in qs
    ]
    return JsonResponse({"client_name": client.name, "daily_stats": daily_stats})


@login_required
def client_dashboard(request):
    """Client-specific dashboard with analytics breakdown by social networks"""
    # Get client based on user role
    client = None
    if request.user.is_superuser:
        client_id = request.GET.get("client_id")
        if client_id:
            client = Client.objects.filter(id=client_id).select_related("agency").first()
    else:
        # Non-superusers can only see their own client data
        client = Client.objects.filter(user=request.user).select_related("agency").first()
    
    if not client:
        return render(request, "cabinet/error.html", {"message": "Client not found or access denied."})
    
    # Get time period parameter
    days_param = request.GET.get("days")
    if days_param:
        try:
            days = int(days_param)
        except ValueError:
            days = None
    else:
        days = None  # All time by default
    
    # Initialize analytics service
    analytics_service = AnalyticsService(client)
    
    # Get analytics data
    hashtag_details = analytics_service.get_hashtag_details()
    network_breakdown = analytics_service.get_network_breakdown(days=days)
    combined_summary = analytics_service.get_combined_analytics_summary(days=days)
    
    # Get client hashtags
    hashtags = list(analytics_service.get_client_hashtags().values_list("hashtag", flat=True))
    
    # Get daily and weekly stats for charts
    daily_stats = analytics_service.get_daily_stats(days=7)
    weekly_stats = analytics_service.get_weekly_stats(weeks=12)
    
    # Convert daily_stats to JSON for JavaScript
    import json
    daily_stats_json = json.dumps(daily_stats)
    
    context = {
        "client": client,
        "hashtags": hashtags,
        "details": hashtag_details,  # For backward compatibility
        "network_breakdown": network_breakdown,
        "combined_summary": combined_summary,
        "daily_stats": daily_stats_json,  # JSON string for JavaScript
        "weekly_stats": weekly_stats,
        "kpi_total_views": combined_summary.total_views,
        "kpi_total_videos": combined_summary.total_posts,
        "kpi_avg_per_video": combined_summary.average_views,
        "kpi_engagement_rate": combined_summary.engagement_rate,
    }
    
    return render(request, "cabinet/client_dashboard.html", context)


@login_required
def agency_clients_view(request):
    agency_id = request.GET.get("agency_id")
    if not agency_id:
        return JsonResponse({"error": "Agency ID required"}, status=400)
    agency = Agency.objects.filter(id=agency_id).first()
    if not agency:
        return JsonResponse({"error": "Agency not found"}, status=404)
    # Authorization: only superuser or the owner of the agency
    if not request.user.is_superuser and agency.owner_id != request.user.id:
        return JsonResponse({"error": "Forbidden"}, status=403)
    clients = agency.clients.all()
    return JsonResponse({
        "agency_name": agency.name,
        "clients": [{"id": c.id, "name": c.name} for c in clients],
    })


@login_required
def dashboard_data_view(request):
    # Restrict counts and stats by role
    if request.user.is_superuser:
        agencies_count = Agency.objects.count()
        clients_count = Client.objects.count()
        tracked_hashtags_qs = ClientHashtag.objects.values_list("hashtag", flat=True).distinct()
        filter_kwargs = {}
    else:
        my_client = Client.objects.filter(user=request.user).first()
        if my_client:
            agencies_count = 1
            clients_count = 1
            tracked_hashtags_qs = ClientHashtag.objects.filter(client=my_client).values_list("hashtag", flat=True).distinct()
            filter_kwargs = {"hashtag__in": list(tracked_hashtags_qs)}
        else:
            my_agency = Agency.objects.filter(owner=request.user).first()
            if my_agency:
                agencies_count = 1
                clients_count = my_agency.clients.count()
                tracked_hashtags_qs = ClientHashtag.objects.filter(client__agency=my_agency).values_list("hashtag", flat=True).distinct()
                filter_kwargs = {"hashtag__in": list(tracked_hashtags_qs)}
            else:
                agencies_count = 0
                clients_count = 0
                tracked_hashtags_qs = ClientHashtag.objects.none()
                filter_kwargs = {"id__isnull": True}
    hashtags_count = len(list(tracked_hashtags_qs))

    end = timezone.now()
    start = end - timezone.timedelta(days=7)
    qs = (
        HashtagAnalytics.objects.filter(created_at__gte=start, created_at__lte=end, **filter_kwargs)
        .annotate(day=TruncDate("created_at"))
        .values("day")
        .annotate(total_views=Sum("total_views"))
        .order_by("day")
    )
    daily_stats = [
        {
            "date": row["day"].strftime("%Y-%m-%d") if row["day"] else "",
            "day_name": row["day"].strftime("%b %d") if row["day"] else "",
            "total_views": int(row["total_views"] or 0),
        }
        for row in qs
    ]
    return JsonResponse({
        "agencies_count": agencies_count,
        "clients_count": clients_count,
        "hashtags_count": hashtags_count,
        "daily_stats": daily_stats,
    })


@login_required
def hashtag_detail(request, hashtag: str):
    hashtag = (hashtag or "").lstrip("#").strip()
    # Access control: superuser sees all; client sees only own; agency sees own clients
    if not request.user.is_superuser:
        my_client = Client.objects.filter(user=request.user).first()
        if my_client:
            if not ClientHashtag.objects.filter(client=my_client, hashtag=hashtag).exists():
                return render(request, "cabinet/error.html", {"message": "Access denied for this hashtag."})
        else:
            my_agency = Agency.objects.filter(owner=request.user).first()
            if my_agency:
                if not ClientHashtag.objects.filter(client__agency=my_agency, hashtag=hashtag).exists():
                    return render(request, "cabinet/error.html", {"message": "Access denied for this hashtag."})
            else:
                return render(request, "cabinet/error.html", {"message": "No cabinet role assigned."})
    qs = HashtagAnalytics.objects.filter(hashtag=hashtag).order_by("created_at")
    # Build daily aggregates
    end = timezone.now()
    start = end - timezone.timedelta(days=30)
    agg = (
        HashtagAnalytics.objects.filter(hashtag=hashtag, created_at__gte=start, created_at__lte=end)
        .annotate(day=TruncDate("created_at"))
        .values("day")
        .annotate(total_views=Sum("total_views"), total_videos=Sum("analyzed_medias"))
        .order_by("day")
    )
    daily_stats = [
        {
            "date": row["day"].strftime("%Y-%m-%d") if row["day"] else "",
            "day_name": row["day"].strftime("%b %d") if row["day"] else "",
            "total_views": int(row.get("total_views") or 0),
            "total_videos": int(row.get("total_videos") or 0),
        }
        for row in agg
    ]
    latest = qs.order_by("-created_at").first()
    return render(
        request,
        "cabinet/hashtag_detail.html",
        {
            "hashtag": hashtag,
            "latest": latest,
            "daily_stats": daily_stats,
        },
    )


@login_required
def account_detail_analytics(request, account_id: int):
    account = get_object_or_404(InstagramAccount, id=account_id)
    # Access control for account ownership
    if not request.user.is_superuser:
        my_client = Client.objects.filter(user=request.user).first()
        if my_client:
            if account.client_id != my_client.id:
                return render(request, "cabinet/error.html", {"message": "Access denied for this account."})
        else:
            my_agency = Agency.objects.filter(owner=request.user).first()
            if my_agency:
                if not account.client or account.client.agency_id != my_agency.id:
                    return render(request, "cabinet/error.html", {"message": "Access denied for this account."})
            else:
                return render(request, "cabinet/error.html", {"message": "No cabinet role assigned."})
    # Last 30 days snapshots if any (we create AccountAnalytics periodically)
    end = timezone.now()
    start = end - timezone.timedelta(days=30)
    agg = (
        AccountAnalytics.objects.filter(account=account, created_at__gte=start, created_at__lte=end)
        .annotate(day=TruncDate("created_at"))
        .values("day")
        .annotate(total_views=Sum("total_views"), total_videos=Sum("total_videos"))
        .order_by("day")
    )
    daily_stats = [
        {
            "date": row["day"].strftime("%Y-%m-%d") if row["day"] else "",
            "day_name": row["day"].strftime("%b %d") if row["day"] else "",
            "total_views": int(row.get("total_views") or 0),
            "total_videos": int(row.get("total_videos") or 0),
        }
        for row in agg
    ]
    latest = AccountAnalytics.objects.filter(account=account).order_by("-created_at").first()
    return render(
        request,
        "cabinet/account_detail.html",
        {
            "account": account,
            "latest": latest,
            "daily_stats": daily_stats,
        },
    )


@login_required
def agency_calculator(request):
    # Only superuser or agency owner
    if not request.user.is_superuser:
        my_agency = Agency.objects.filter(owner=request.user).first()
        if not my_agency:
            return redirect("cabinet_dashboard")
    return render(request, "cabinet/agency_calculator.html", {})


@login_required
@require_http_methods(["POST"])
def agency_calc_quote(request):
    # Only superuser or agency owner
    user_agency = None
    if not request.user.is_superuser:
        user_agency = Agency.objects.filter(owner=request.user).first()
        if not user_agency:
            return JsonResponse({"error": "Forbidden"}, status=403)
    
    try:
        data = json.loads(request.body.decode("utf-8")) if request.body else {}
    except Exception:
        data = {}
    
    service = AgencyCalculatorService()
    result = service.calculate(data)
    
    if "error" not in result:
        # Сохраняем расчет в базу данных
        try:
            inputs = result.get("inputs", {})
            totals = result.get("totals", {})
            percents = result.get("percents", {})
            
            # Конвертируем в USD для сохранения
            usd_to_rub = 92.5
            final_cost_usd = totals.get("rub_total", 0) / usd_to_rub
            
            # Подхватываем имя клиента, если пришло в payload (после модального окна)
            client_name = (data.get("client_name") or "").strip()

            calculation = CalculationHistory.objects.create(
                user=request.user,
                agency=user_agency,
                client_name=client_name,
                volume_millions=inputs.get("volume_millions", 0),
                platforms=inputs.get("platforms", []),
                countries=inputs.get("countries", []),
                currency=data.get("currency", "RUB"),
                own_badge=inputs.get("flags", {}).get("own_badge", False),
                own_content=inputs.get("flags", {}).get("own_content", False),
                pilot=inputs.get("flags", {}).get("pilot", False),
                vip_percent=inputs.get("vip_percent", 0),
                urgent=inputs.get("flags", {}).get("urgent", False),
                peak_season=inputs.get("flags", {}).get("peak_season", False),
                exclusive_content=inputs.get("flags", {}).get("exclusive_content", False),
                base_price_per_view=result.get("unit_prices", {}).get("rub_per_view", 0),
                tier_multiplier=result.get("tier", {}).get("multiplier", 1.0),
                platform_multiplier=result.get("platform_multiplier", 1.0),
                discounts_percent=percents.get("discounts", 0),
                surcharges_percent=percents.get("surcharges", 0),
                market_discount_percent=percents.get("market_discount", 0),
                final_cost_rub=totals.get("rub_total", 0),
                final_cost_usd=final_cost_usd,
                calculation_data=result,
                notes=data.get("notes", "")
            )
            result["calculation_id"] = calculation.id
            result["saved"] = True
        except Exception as e:
            result["save_error"] = str(e)
    
    status = 200 if "error" not in result else 400
    return JsonResponse(result, status=status)


@login_required
def export_calculations_csv(request):
    """Export calculation history to CSV"""
    user = request.user
    
    # Получаем расчеты в зависимости от роли
    if user.is_superuser:
        calculations = CalculationHistory.objects.all()
        filename = f"all_calculations_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
    else:
        my_agency = Agency.objects.filter(owner=user).first()
        if my_agency:
            calculations = CalculationHistory.objects.filter(agency=my_agency)
            filename = f"agency_{my_agency.name}_calculations_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
        else:
            calculations = CalculationHistory.objects.filter(user=user)
            filename = f"user_calculations_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
    
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Добавляем BOM для корректного отображения кириллицы в Excel
    response.write('\ufeff')
    
    writer = csv.writer(response)
    
    # Заголовки
    writer.writerow([
        'ID',
        'Дата создания',
        'Пользователь',
        'Агентство',
        'Имя клиента',
        'Объем (млн)',
        'Платформы',
        'Страны',
        'Валюта',
        'Базовая цена за просмотр',
        'Множитель по стране',
        'Множитель платформы',
        'Скидки (%)',
        'Надбавки (%)',
        'Рыночная скидка (%)',
        'Итого RUB',
        'Итого USD',
        'Свой бейдж',
        'Свой контент',
        'Пилотный проект',
        'VIP скидка (%)',
        'Срочность',
        'Пиковый сезон',
        'Эксклюзивный контент',
        'Заметки'
    ])
    
    # Данные
    for calc in calculations.select_related('user', 'agency'):
        platforms_str = ', '.join(calc.platforms) if calc.platforms else '-'
        countries_str = ', '.join([c.get('code', '') for c in calc.countries]) if calc.countries else '-'
        
        writer.writerow([
            calc.id,
            calc.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            calc.user.username if calc.user else '-',
            calc.agency.name if calc.agency else '-',
            calc.client_name or '-',
            calc.volume_millions,
            platforms_str,
            countries_str,
            calc.currency,
            f"{calc.base_price_per_view:.6f}",
            f"{calc.tier_multiplier:.3f}",
            f"{calc.platform_multiplier:.3f}",
            f"{calc.discounts_percent:.1f}",
            f"{calc.surcharges_percent:.1f}",
            f"{calc.market_discount_percent:.1f}",
            f"{calc.final_cost_rub:.2f}",
            f"{calc.final_cost_usd:.2f}",
            'Да' if calc.own_badge else 'Нет',
            'Да' if calc.own_content else 'Нет',
            'Да' if calc.pilot else 'Нет',
            f"{calc.vip_percent * 100:.1f}",
            'Да' if calc.urgent else 'Нет',
            'Да' if calc.peak_season else 'Нет',
            'Да' if calc.exclusive_content else 'Нет',
            calc.notes or '-'
        ])
    
    return response


@login_required
def download_calculation_excel(request, calculation_id):
    """Download specific calculation as Excel"""
    calculation = CalculationHistory.objects.filter(id=calculation_id).first()
    if not calculation:
        return HttpResponse("Расчет не найден", status=404)
    
    # Проверка доступа
    if not request.user.is_superuser:
        if calculation.user != request.user:
            my_agency = Agency.objects.filter(owner=request.user).first()
            if not my_agency or calculation.agency != my_agency:
                return HttpResponse("Доступ запрещен", status=403)
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Расчет {calculation.id}"
    
    # Заголовок
    ws.append([f"Расчет стоимости продвижения #{calculation.id}"])
    ws.append([f"Дата: {calculation.created_at.strftime('%Y-%m-%d %H:%M:%S')}"])
    ws.append([])
    
    # Входные параметры
    ws.append(["ВХОДНЫЕ ПАРАМЕТРЫ"])
    ws.append(["Объем просмотров (млн):", calculation.volume_millions])
    ws.append(["Платформы:", ', '.join(calculation.platforms)])
    ws.append(["Страны:", ', '.join([c.get('code', '') for c in calculation.countries])])
    ws.append(["Валюта:", calculation.currency])
    ws.append([])
    
    # Скидки и надбавки
    ws.append(["СКИДКИ И НАДБАВКИ"])
    ws.append(["Свой бейдж:", "Да" if calculation.own_badge else "Нет"])
    ws.append(["Свой контент:", "Да" if calculation.own_content else "Нет"])
    ws.append(["Пилотный проект:", "Да" if calculation.pilot else "Нет"])
    ws.append(["VIP скидка (%):", f"{calculation.vip_percent * 100:.1f}"])
    ws.append(["Срочность:", "Да" if calculation.urgent else "Нет"])
    ws.append(["Пиковый сезон:", "Да" if calculation.peak_season else "Нет"])
    ws.append(["Эксклюзивный контент:", "Да" if calculation.exclusive_content else "Нет"])
    ws.append([])
    
    # Результаты расчета
    ws.append(["РЕЗУЛЬТАТЫ РАСЧЕТА"])
    ws.append(["Базовая цена за просмотр:", f"{calculation.base_price_per_view:.6f} ₽"])
    ws.append(["Множитель по стране:", f"{calculation.tier_multiplier:.3f}"])
    ws.append(["Множитель платформы:", f"{calculation.platform_multiplier:.3f}"])
    ws.append(["Скидки (%):", f"{calculation.discounts_percent:.1f}"])
    ws.append(["Надбавки (%):", f"{calculation.surcharges_percent:.1f}"])
    ws.append(["Рыночная скидка (%):", f"{calculation.market_discount_percent:.1f}"])
    ws.append([])
    ws.append(["ИТОГОВАЯ СТОИМОСТЬ"])
    ws.append(["В рублях:", f"{calculation.final_cost_rub:.2f} ₽"])
    ws.append(["В долларах:", f"{calculation.final_cost_usd:.2f} $"])
    
    if calculation.notes:
        ws.append([])
        ws.append(["ЗАМЕТКИ"])
        ws.append([calculation.notes])
    
    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    filename = f"calculation_{calculation.id}_{calculation.created_at.strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    response = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f"attachment; filename={filename}"
    return response


@login_required
def calculation_details_api(request, calculation_id):
    """API endpoint to get calculation details"""
    calculation = CalculationHistory.objects.filter(id=calculation_id).first()
    if not calculation:
        return JsonResponse({"success": False, "error": "Calculation not found"}, status=404)
    
    # Check access
    if not request.user.is_superuser:
        if calculation.user != request.user:
            my_agency = Agency.objects.filter(owner=request.user).first()
            if not my_agency or calculation.agency != my_agency:
                return JsonResponse({"success": False, "error": "Access denied"}, status=403)
    
    data = {
        "success": True,
        "calculation": {
            "id": calculation.id,
            "client_name": calculation.client_name,
            "volume_millions": calculation.volume_millions,
            "platforms": calculation.platforms,
            "countries": calculation.countries,
            "currency": calculation.currency,
            "own_badge": calculation.own_badge,
            "own_content": calculation.own_content,
            "pilot": calculation.pilot,
            "vip_percent": calculation.vip_percent,
            "urgent": calculation.urgent,
            "peak_season": calculation.peak_season,
            "exclusive_content": calculation.exclusive_content,
            "base_price_per_view": calculation.base_price_per_view,
            "tier_multiplier": calculation.tier_multiplier,
            "platform_multiplier": calculation.platform_multiplier,
            "discounts_percent": calculation.discounts_percent,
            "surcharges_percent": calculation.surcharges_percent,
            "market_discount_percent": calculation.market_discount_percent,
            "final_cost_rub": calculation.final_cost_rub,
            "final_cost_usd": calculation.final_cost_usd,
            "notes": calculation.notes,
            "created_at": calculation.created_at.isoformat(),
            "user_name": calculation.user.username if calculation.user else "Unknown",
        }
    }
    
    return JsonResponse(data)


@login_required
def calculations_list_api(request):
    """API endpoint to get more calculations with pagination"""
    offset = int(request.GET.get('offset', 0))
    limit = 10
    
    user = request.user
    
    # Filter calculations based on user role
    if user.is_superuser:
        calculations = CalculationHistory.objects.all()
    else:
        my_agency = Agency.objects.filter(owner=user).first()
        if my_agency:
            calculations = CalculationHistory.objects.filter(agency=my_agency)
        else:
            calculations = CalculationHistory.objects.filter(user=user)
    
    calculations = calculations.select_related('user', 'agency').order_by('-created_at')[offset:offset+limit]
    
    data = {
        "success": True,
        "calculations": [
            {
                "id": calc.id,
                "client_name": calc.client_name,
                "volume_millions": calc.volume_millions,
                "platforms": calc.platforms,
                "countries": calc.countries,
                "currency": calc.currency,
                "final_cost_rub": calc.final_cost_rub,
                "final_cost_usd": calc.final_cost_usd,
                "created_at": calc.created_at.isoformat(),
                "user_name": calc.user.username if calc.user else "Unknown",
            }
            for calc in calculations
        ]
    }
    
    return JsonResponse(data)


@login_required
def admin_analytics_api(request):
    """API endpoint for filtered admin analytics"""
    if not request.user.is_superuser:
        return JsonResponse({"success": False, "error": "Forbidden"}, status=403)
    
    agency_id = request.GET.get('agency_id')
    client_id = request.GET.get('client_id')
    days = int(request.GET.get('days', 7))
    
    # Time range
    end = timezone.now()
    start = end - timezone.timedelta(days=days)
    
    # Build filter conditions
    hashtag_filter = {}
    if agency_id:
        agency = Agency.objects.filter(id=agency_id).first()
        if agency:
            agency_clients = list(agency.clients.all())
            agency_hashtags = ClientHashtag.objects.filter(client__in=agency_clients).values_list("hashtag", flat=True)
            hashtag_filter["hashtag__in"] = list(agency_hashtags)
    elif client_id:
        client = Client.objects.filter(id=client_id).first()
        if client:
            client_hashtags = ClientHashtag.objects.filter(client=client).values_list("hashtag", flat=True)
            hashtag_filter["hashtag__in"] = list(client_hashtags)
    
    # Daily stats
    qs = (
        HashtagAnalytics.objects.filter(created_at__gte=start, created_at__lte=end, **hashtag_filter)
        .annotate(day=TruncDate("created_at"))
        .values("day")
        .annotate(
            total_views=Sum("total_views"),
            total_videos=Sum("analyzed_medias"),
            total_likes=Sum("total_likes"),
            total_comments=Sum("total_comments"),
        )
        .order_by("day")
    )
    daily_stats = [
        {
            "date": row["day"].strftime("%Y-%m-%d") if row["day"] else "",
            "day_name": row["day"].strftime("%b %d") if row["day"] else "",
            "total_views": int(row.get("total_views") or 0),
            "total_videos": int(row.get("total_videos") or 0),
            "total_likes": int(row.get("total_likes") or 0),
            "total_comments": int(row.get("total_comments") or 0),
        }
        for row in qs
    ]
    
    # Agencies analytics
    agencies_rows = []
    if not client_id:  # Only show agencies if not filtering by specific client
        agencies = Agency.objects.all()
        for a in agencies:
            a_clients = list(a.clients.all())
            a_hashtags = ClientHashtag.objects.filter(client__in=a_clients).values_list("hashtag", flat=True)
            if not a_hashtags:
                continue
            agg = (
                HashtagAnalytics.objects.filter(hashtag__in=list(a_hashtags), created_at__gte=start, created_at__lte=end)
                .aggregate(
                    views=Sum("total_views"),
                    videos=Sum("analyzed_medias"),
                )
            )
            views_sum = int(agg.get("views") or 0)
            videos_sum = int(agg.get("videos") or 0)
            if views_sum == 0 and videos_sum == 0:
                continue
            avg_views = (views_sum / videos_sum) if videos_sum > 0 else 0.0
            agencies_rows.append({
                "agency": {"id": a.id, "name": a.name},
                "views": views_sum,
                "videos": videos_sum,
                "avg_views": avg_views,
            })
        agencies_rows = sorted(agencies_rows, key=lambda x: x["views"], reverse=True)
    
    # Clients analytics
    clients_rows = []
    clients = Client.objects.all()
    if agency_id:
        clients = clients.filter(agency_id=agency_id)
    elif client_id:
        clients = clients.filter(id=client_id)
    
    for c in clients:
        c_hashtags = ClientHashtag.objects.filter(client=c).values_list("hashtag", flat=True)
        if not c_hashtags:
            continue
        agg = (
            HashtagAnalytics.objects.filter(hashtag__in=list(c_hashtags), created_at__gte=start, created_at__lte=end)
            .aggregate(
                views=Sum("total_views"),
                videos=Sum("analyzed_medias"),
            )
        )
        views_sum = int(agg.get("views") or 0)
        videos_sum = int(agg.get("videos") or 0)
        if views_sum == 0 and videos_sum == 0:
            continue
        avg_views = (views_sum / videos_sum) if videos_sum > 0 else 0.0
        clients_rows.append({
            "client": {"id": c.id, "name": c.name},
            "agency": {"id": c.agency.id, "name": c.agency.name} if c.agency else None,
            "views": views_sum,
            "videos": videos_sum,
            "avg_views": avg_views,
        })
    clients_rows = sorted(clients_rows, key=lambda x: x["views"], reverse=True)
    
    # Top hashtags
    hashtags_breakdown = []
    if not hashtag_filter:  # Only show all hashtags if not filtering
        hashtag_qs = (
            HashtagAnalytics.objects.filter(created_at__gte=start, created_at__lte=end)
            .values("hashtag")
            .annotate(
                views=Sum("total_views"),
                videos=Sum("analyzed_medias"),
                likes=Sum("total_likes"),
                comments=Sum("total_comments"),
            )
            .order_by("-views")[:20]
        )
        hashtags_breakdown = [
            {
                "hashtag": row["hashtag"],
                "total_views": int(row.get("views") or 0),
                "analyzed_medias": int(row.get("videos") or 0),
                "total_likes": int(row.get("likes") or 0),
                "total_comments": int(row.get("comments") or 0),
            }
            for row in hashtag_qs
        ]
    
    return JsonResponse({
        "success": True,
        "dailyStats": daily_stats,
        "agenciesRows": agencies_rows,
        "clientsRows": clients_rows,
        "hashtagsBreakdown": hashtags_breakdown,
    })


# === Currency API endpoints ===

@login_required
@require_http_methods(["GET"])
def get_currency_rates(request):
    """
    Получить актуальные курсы валют
    """
    # Только superuser или владелец агентства
    if not request.user.is_superuser:
        my_agency = Agency.objects.filter(owner=request.user).first()
        if not my_agency:
            return JsonResponse({"error": "Forbidden"}, status=403)
    
    try:
        rates_info = currency_service.get_rates_info()
        return JsonResponse({
            "success": True,
            "rates": rates_info,
            "timestamp": timezone.now().isoformat(),
        })
    except Exception as e:
        return JsonResponse({
            "success": False,
            "error": str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def force_update_currency_rates(request):
    """
    Принудительно обновить курсы валют
    """
    # Только superuser
    if not request.user.is_superuser:
        return JsonResponse({"error": "Forbidden"}, status=403)
    
    try:
        rates = currency_service.force_update()
        return JsonResponse({
            "success": True,
            "message": "Курсы валют обновлены",
            "rates": rates,
            "timestamp": timezone.now().isoformat(),
        })
    except Exception as e:
        return JsonResponse({
            "success": False,
            "error": str(e)
        }, status=500)
